from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from tkinter import *
import tkinter as tk
import numpy as np
import time
from PIL import Image,ImageTk
from matplotlib import style
from time import strftime
import cv2
from math import atan2, cos, sin, sqrt, pi
import openpyxl
import pandas as pd
import math
cond = False
# area_n = 0
# --- Deklarasi Variabel ---
style.use("dark_background")

fc1="#171F24"
bg1="#03DAC5"
bg2="#CF6679"
gold = "#ffd700"
fg1="#FFFFFF"
fg2="#000000"
text1="#FFFFFF"
fill2="#171F24"

FontGede=("Gama-Sans", 16)
FontCilik=("Gama-Serif", 12)
FontCilikBgt=("Gama-Serif",13)

cap1 = cv2.VideoCapture(0)
cap1.set(cv2.CAP_PROP_FRAME_WIDTH,400)
cap1.set(cv2.CAP_PROP_FRAME_HEIGHT,200)

# cap2 = cv2.VideoCapture(2)
# cap2.set(cv2.CAP_PROP_FRAME_WIDTH,400)
# cap2.set(cv2.CAP_PROP_FRAME_HEIGHT,200)

width_cm = 255.0 / 640
height_cm = 300.0 / 480

init_val = [0,108,122,193,85,183]

lower_green = np.array([0,122, 85])
upper_green = np.array([108,193, 183])

lower_red = np.array([0,12,131])
upper_red = np.array([51,158,186])
# --- Function ----
# def time_real():
#     string = strftime('%H:%M:%S %p')
#     lbl.config(text=string)
#     lbl.after(1000, time)

def color_detection(img, lower, upper, id):
  x_cm = 0
  y_cm = 0
  center = 0
  try:
    mask = cv2.inRange(img, lower, upper)
    hsl_result = cv2.bitwise_and(img, img, mask = mask)

    gray = cv2.cvtColor(hsl_result, cv2.COLOR_BGR2GRAY)
      
    _, bw = cv2.threshold(gray, 50, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
      
    contours, _ = cv2.findContours(bw, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
    # id_val = id  
    for i, c in enumerate(contours):
      area = cv2.contourArea(c)
      if area < 90 or 100000 < area:
        continue
      peri = cv2.arcLength(c, True)
      approx = cv2.approxPolyDP(c, 0.02 * peri, True)
      #print(len(approx))
      x , y , w, h = cv2.boundingRect(approx)

      x_center = x+int(w//2)
      y_center = y+int((h//2))

      # x_cm = x+int(w//2)
      # y_cm = y+int((h//2))
      center = (x_center,y_center)
      cv2.circle(img,center,4,(0,255,0),-1)
      x_cm = (x_center / 2.710417)
      y_cm = (y_center / 1.820833) #Adjust from frame and real
      cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)
      
      text_loct = (x+(w//2)-10,y+(h//2)-26)
      # degree = getOrientation(c, img)
      cv2.putText(img,str(id),text_loct,cv2.FONT_HERSHEY_COMPLEX,0.9,(0,0,255),2)
      # cv2.putText(img,"+",center,cv2.FONT_HERSHEY_COMPLEX,0.9,(0,255,255),2)
      # cv2.drawContours(img, contours, i, (0, 0, 255), 2)
      return x_cm, y_cm, center
    
  except:
    return x_cm, y_cm, center

def open_camera():
  global img
  global x_cm
  global y_cm
  global degree
  global start_time
  global end_time
  global id
  # last_degrees = 0
  # _, img = cap1.read()

  start_time = time.time()
  success1, frame1 = cap1.read()
  success2, frame2 = cap2.read()
  m_frame1 = cv2.flip(frame1, 1)
  m_frame2 = cv2.flip(frame2, 1)
  r_frame2 = cv2.rotate(m_frame2, cv2.ROTATE_180)
  img_1 = np.concatenate((r_frame2, m_frame1), axis=0)
  img_2 = cv2.resize(img_1, (640,480))
  img_3 = cv2.rotate(img_2, cv2.ROTATE_180)
  img = cv2.flip(img_3, 1)

  try:
    x_green, y_green, center_green = color_detection(img, lower_green, upper_green, id=1)
    x_red, y_red, center_red = color_detection(img, lower_red, upper_red, id=2)
    
    jarak_x = (x_red - x_green)*(x_red - x_green)
    jarak_y = (y_red - y_green)* (y_red - y_green)
    line_length = sqrt(jarak_x + jarak_y)

    center_x = (x_red + x_green)/2
    center_y = (y_red + y_green)/2
    coor_line = (center_x, center_y) # Titik pusat

    delta_x = x_red - x_green
    delta_y = y_red - y_green

    angle_radians = math.atan2(delta_y, delta_x)
    angle_degrees = math.degrees(angle_radians)

    if angle_degrees < 0:
      angle_degrees = angle_degrees + 360

    cv2.arrowedLine(img, center_green, center_red, (255, 255, 0),5)
    Label(canvas, font=("Gama-Sans",18, "bold"),background=gold, foreground=fg2 ,text="                  ").place(x=400,y=120)
    Label(canvas, font=("Gama-Sans",18, "bold"),background=gold, foreground=fg2 ,text="Degree : " + str(int(angle_degrees))).place(x=400,y=120)

    # if angle_degrees != last_degrees:
    #   label_text.destroy()

    # last_degrees = angle_degrees
    Label(canvas, font=("Gama-Sans",18, "bold"),background=gold, foreground=fg2 ,text="                   ").place(x=100,y=120)
    Label(canvas, font=("Gama-Sans",18, "bold"),background=gold, foreground=fg2 ,text="(X,Y) : (" + str(int(center_x))+", " +str(int(center_y))+")").place(x=100,y=120)

    if(cond==True):
      write_excel(center_x, center_y, angle_degrees)
      
    # Convert image from one color space to other
    end_time = time.time()
    milliseconds = (end_time - start_time)*10**3
    text_time = Label(canvas, font=("Gama-Sans",18, "bold"),background=gold, foreground=fg2 ,text="Sampling Time : " + str(milliseconds))
    text_time.place(x=700,y=120)

  except:
    pass


  opencv_image = cv2.cvtColor(img, cv2.COLOR_BGR2RGBA)
  
    # Capture the latest frame and transform to image
  captured_image = Image.fromarray(opencv_image)
  
    # Convert captured image to photoimage
  photo_image = ImageTk.PhotoImage(image=captured_image)
  
    # Displaying photoimage in the label
  label_widget.photo_image = photo_image
  
    # Configure image in the label
  label_widget.configure(image=photo_image)
  
    # Repeat the same process after every 15 seconds
  label_widget.after(15, open_camera)


def plot_start():
  global cond
  cond = True

def plot_stop():
  global cond
  cond = False

def drawAxis(img, p_, q_, color, scale):
  p = list(p_)
  q = list(q_)
 
  ## [visualization1]
  angle = atan2(p[1] - q[1], p[0] - q[0]) # angle in radians
  hypotenuse = sqrt((p[1] - q[1]) * (p[1] - q[1]) + (p[0] - q[0]) * (p[0] - q[0]))
 
  # Here we lengthen the arrow by a factor of scale
  q[0] = p[0] - scale * hypotenuse * cos(angle)
  q[1] = p[1] - scale * hypotenuse * sin(angle)
  cv2.line(img, (int(p[0]), int(p[1])), (int(q[0]), int(q[1])), color, 3, cv2.LINE_AA)
 
  # create the arrow hooks
  p[0] = q[0] + 9 * cos(angle + pi / 4)
  p[1] = q[1] + 9 * sin(angle + pi / 4)
  cv2.line(img, (int(p[0]), int(p[1])), (int(q[0]), int(q[1])), color, 3, cv2.LINE_AA)
 
  p[0] = q[0] + 9 * cos(angle - pi / 4)
  p[1] = q[1] + 9 * sin(angle - pi / 4)
  cv2.line(img, (int(p[0]), int(p[1])), (int(q[0]), int(q[1])), color, 3, cv2.LINE_AA)
  ## [visualization1]
 
def getOrientation(pts, img):
  ## [pca]
  # Construct a buffer used by the pca analysis
  sz = len(pts)
  data_pts = np.empty((sz, 2), dtype=np.float64)
  for i in range(data_pts.shape[0]):
    data_pts[i,0] = pts[i,0,0]
    data_pts[i,1] = pts[i,0,1]
 
  # Perform PCA analysis
  mean = np.empty((0))
  mean, eigenvectors, eigenvalues = cv2.PCACompute2(data_pts, mean)
 
  # Store the center of the object
  cntr = (int(mean[0,0]), int(mean[0,1]))
  ## [pca]
 
  ## [visualization]
  # Draw the principal components
  cv2.circle(img, cntr, 3, (255, 0, 255), 2)
  p1 = (cntr[0] + 0.02 * eigenvectors[0,0] * eigenvalues[0,0], cntr[1] + 0.02 * eigenvectors[0,1] * eigenvalues[0,0])
  p2 = (cntr[0] - 0.02 * eigenvectors[1,0] * eigenvalues[1,0], cntr[1] - 0.02 * eigenvectors[1,1] * eigenvalues[1,0])
  drawAxis(img, cntr, p1, (255, 255, 0), 5)
  # drawAxis(img, cntr, p2, (0, 0, 255), 5)
 
  angle = atan2(eigenvectors[0,1], eigenvectors[0,0]) # orientation in radians
  ## [visualization]
 
  # Label with the rotation angle
  angle = (int(np.rad2deg(angle)) - 90)*-1
  # label = "  Rotation Angle: " + str(angle) + " degrees"
  # textbox = cv2.rectangle(img, (cntr[0], cntr[1]-25), (cntr[0] + 250, cntr[1] + 10), (255,255,255), -1)
  # cv2.putText(img, label, (cntr[0], cntr[1]), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,0,0), 1, cv2.LINE_AA)
 
  return angle

def plot_data():
  try:
    file_path = 'Data_testbed.xlsx'

    df = pd.read_excel(file_path)

    # Load the Excel file into a DataFrame
    # workbook = openpyxl.load_workbook(file_path)
    # sheet = workbook['Sheet1']

    # for val in sheet.iter_rows(max_col=3):
    #       coor_x = val[0].value
    #       coor_y = val[1].value
    coor_x = df['X']
    coor_y = df['Y']
    # print(coor_x)

    u = np.diff(coor_x)
    v = np.diff(coor_y)
    pos_x = coor_x[:-1] + u/2
    pos_y = coor_y[:-1] + v/2
    norm = np.sqrt(u**2+v**2) 

    fig3 = Figure()

    ax3 = fig3.add_subplot(111, facecolor = gold)

    ax3.plot(coor_x,coor_y, marker="o")
    ax3.quiver(pos_x, pos_y, u/norm, v/norm, angles="xy", zorder=5, pivot="mid")
    ax3.grid()
    # ax3.legend()
    ax3.set_title("Posisi koordinat Robot")
    ax3.set_xlabel("X")
    ax3.set_ylabel("Y")
    ax3.set_xlim(0,300)
    ax3.set_ylim(0,300)
    ax3.invert_yaxis()

    canvas3 = FigureCanvasTkAgg(fig3, master=root)
    canvas3.get_tk_widget().place(x=850,y=180,width=600,height=490)
    canvas3.draw()
    
    root.update()

  except:
     pass
  # root.after(10, plot_data)

def nothing(a):
    pass

def write_excel(data1, data2, data3):
    xlsx_file_path = 'Data_testbed.xlsx'

    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active

    new_row = [data1, data2, data3]

    sheet.append(new_row)

    workbook.save('Data_testbed.xlsx')

    # Close the workbook
    workbook.close()

def delete_excel():
    xlsx_file_path = 'Data_testbed.xlsx'

    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active

    sheet.delete_cols(1,3)

    new_header = ['X', 'Y', 'Degree']

    sheet.append(new_header)

    workbook.save('Data_testbed.xlsx')

    # Close the workbook
    workbook.close()

  
# ---- Setup GUI ----
root = tk.Tk()
root.title('Real Time Plot')

root.geometry("1920x1080")
# root.attributes('-fullscreen', 1)
root.bind('<Escape>', lambda _: root.destroy())

bg = PhotoImage(file = "5053309.png")
canvas = Canvas(root, width=1920, height=1080)
canvas.pack(fill = "both", expand = True)
canvas.create_image( 0, 0, image = bg, anchor = "nw")

judul = "Testbed Orientasi Mobile Robot"

canvas.pack()
canvas.create_text(620, 30, anchor=W, font=("Gama-Sans", 24),text = judul, fill = text1)

lbl = Label(canvas, font=("Gama-Sans",24, "bold"), background=gold, foreground='white')
lbl.pack(anchor=NE)

img_l = Image.open("Logo3.png")
img_l = img_l.resize((200,47), Image.LANCZOS)
logo= ImageTk.PhotoImage(img_l)
canvas.create_image(50,45,anchor=W,image=logo)

# time_real()

# ---- Camera ----

label_widget = Label(canvas) # -- Widget Camera
label_widget.place(x=100, y=180)

open_camera()

# Button
button1 = Button(canvas, text="RECORD", font=("Gama-Sans",13, "bold"), background = gold, activebackground=fill2, foreground= fg2, activeforeground= fg2, width=20,command=plot_start)
button1.place(x=100, y=700)

root.update()

button2 = Button(canvas, text="STOP", font=("Gama-Sans",13, "bold"), background = gold, activebackground=fill2, foreground= fg2, activeforeground= fg2, width=20,command=plot_stop)
button2.place(x=500, y=700)

root.update()

button3 = Button(canvas, text="SHOW", font=("Gama-Sans",13, "bold"), background = gold, activebackground=fill2, foreground= fg2, activeforeground= fg2, width=20,command=plot_data)
button3.place(x=850, y=700)

root.update()

button4 = Button(canvas, text="RESET", font=("Gama-Sans",13, "bold"), background = gold, activebackground=fill2, foreground= fg2, activeforeground= fg2, width=20,command=delete_excel)
button4.place(x=1250, y=700)

root.update()

root.mainloop()